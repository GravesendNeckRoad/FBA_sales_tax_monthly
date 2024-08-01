# ____________________________________________________________________________________________________________________
# This script automates tax files for Johns accountant. It requires a .txt file with the raw sales for the given month.
# Outputs a formatted .xlsx file with a breakdown of revenue/tax by state for the account for that given month.
# ____________________________________________________________________________________________________________________
import os
from datetime import datetime
import pandas as pd
import openpyxl as xl
from openpyxl import utils, styles
import logging

logging.basicConfig(level=logging.ERROR, format='%(levelname)s:\n%(message)s')
pd.set_option('display.width', None)
tax_files_dir = os.path.dirname(__file__)


def main():
    # _________________________________________________________________________________________________________________
    # DEFINING STATIC VARIABLES - ACCOUNT NAME, RAW DATA DIRECTORY, AND DATE-RANGE. CREATE STATE DICTIONARY
    # _________________________________________________________________________________________________________________

    account_name = input("Please enter the name of the account you are running this report for.\n").title()

    file_name = input("Enter the name of your .txt monthly sales file, whilst making sure it is located in your "
                      "tax report folder. Do not use any paths, or file extensions.\n"
                      "Correct example: '540640019814'\n"
                      r"Incorrect example: '540640019814.txt', or '...\Downloads\540640019814.txt'"
                      "\n")

    year_month = input("Please enter the year and month the report is being ran, in the format : YYYY-MM.\n"
                       "Correct example: '2024-03'\n")

    # suppressing warnings (these will be handled & script will terminate if any are left undefined)
    year_month_properly_formatted, total_revenue, total_tax = None, None, None

    try:
        year_month_properly_formatted = datetime.strptime(year_month, '%Y-%m').strftime('%B %Y')
    except (ValueError, Exception) as e:
        logging.error(f'Please enter a valid data in the format %Y-%m\n{e}')
        exit()

    # the raw data has a mix of state abbreviations and names, must clean it up to only show full state names
    states = {
        'ALABAMA': 'AL',
        'ALASKA': 'AK',
        'ARIZONA': 'AZ',
        'ARKANSAS': 'AR',
        'CALIFORNIA': 'CA',
        'COLORADO': 'CO',
        'CONNECTICUT': 'CT',
        'DELAWARE': 'DE',
        'DISTRICT OF COLUMBIA': 'DC',
        'FLORIDA': 'FL',
        'GEORGIA': 'GA',
        'HAWAII': 'HI',
        'IDAHO': 'ID',
        'ILLINOIS': 'IL',
        'INDIANA': 'IN',
        'IOWA': 'IA',
        'KANSAS': 'KS',
        'KENTUCKY': 'KY',
        'LOUISIANA': 'LA',
        'MAINE': 'ME',
        'MARYLAND': 'MD',
        'MASSACHUSETTS': 'MA',
        'MICHIGAN': 'MI',
        'MINNESOTA': 'MN',
        'MISSISSIPPI': 'MS',
        'MISSOURI': 'MO',
        'MONTANA': 'MT',
        'NEBRASKA': 'NE',
        'NEVADA': 'NV',
        'NEW HAMPSHIRE': 'NH',
        'NEW JERSEY': 'NJ',
        'NEW MEXICO': 'NM',
        'NEW YORK': 'NY',
        'NORTH CAROLINA': 'NC',
        'NORTH DAKOTA': 'ND',
        'OHIO': 'OH',
        'OKLAHOMA': 'OK',
        'OREGON': 'OR',
        'PENNSYLVANIA': 'PA',
        'RHODE ISLAND': 'RI',
        'SOUTH CAROLINA': 'SC',
        'SOUTH DAKOTA': 'SD',
        'TENNESSEE': 'TN',
        'TEXAS': 'TX',
        'UTAH': 'UT',
        'VERMONT': 'VT',
        'VIRGINIA': 'VA',
        'WASHINGTON': 'WA',
        'WEST VIRGINIA': 'WV',
        'WISCONSIN': 'WI',
        'WYOMING': 'WY'
    }

    # invert dict to be able to work both ways
    states_inverted = {}
    for keys, values in states.items():
        states_inverted[values] = keys
    # _________________________________________________________________________________________________________________
    # LOAD IN SALES DATA AND WRANGLE IT AS NECESSARY
    # _________________________________________________________________________________________________________________

    # load in the sales data .txt file
    df = pd.DataFrame([])
    for file in os.listdir(tax_files_dir):
        try:
            if file == f'{file_name}.txt':
                df = pd.read_csv(
                    f"{os.path.join(tax_files_dir, file_name)}.txt"
                    .replace("\\", "/")
                    , encoding='latin1'
                    , delimiter='\t'
                    , usecols=[
                        'amazon-order-id'
                        , 'merchant-order-id'
                        , 'item-price'
                        , 'item-tax'
                        , 'ship-state'
                        , 'product-name'
                        , 'purchase-date'
                        , 'ship-country'
                        , 'item-status'
                    ]
                    , dtype={
                        'item-price': 'float32'
                        , 'item-tax': 'float32'
                    }
                )
        except Exception as e:
            logging.error(f"Please inspect your data and try again.\n{e}")
            exit()

    if len(df) == 0:
        logging.exception('No data was read in. Please check your txt sales data feed and try again.')
        exit()

    # wrangling; applying filters and formatting the 'state' column (cleaning incorrect/inconsistent state names)
    # if a row doesnt match up to a state name but state=US, it must be an overseas army territory; replacing w. 'OTHER'
    try:
        df = df.drop_duplicates() \
            .loc[
                lambda x:
                    (x['ship-country'] == 'US') &
                    (x['item-status'] != 'Cancelled') &
                    (x['product-name'] != '-') &
                    (x['purchase-date'].str.startswith(f'{year_month}'))
        ] \
            .assign(
                ship_state_FIXED=lambda x: x['ship-state'].str.upper().replace(states_inverted)
                    .map(lambda y: y if y in states_inverted.values() else 'OTHER')
        ) \
            .fillna(0) \
            .reset_index()

        # perform aggregation by state
        pivot_totals = df.groupby('ship_state_FIXED').agg({'item-price': 'sum', 'item-tax': 'sum'}) \
                .round(2).astype('str').astype('float') \
                .reset_index()

        # merge back to full dictionary (need all 50 states returned even if they have $0 revenue)
        # left join to leave out any OTHER aka overseas territories, could do 'outer' to return territories as OTHER
        df = pd.merge(
            pd.DataFrame(states.keys()),
            pivot_totals,
            left_on=0,
            right_on='ship_state_FIXED',
            how='left'
        ) \
            .fillna(0) \
            .rename({0: 'States', 'item-price': 'Revenue', 'item-tax': 'Tax'}, axis=1) \
            [['States', 'Revenue', 'Tax']]

        # clean up state names (caps lock = ugly)
        df['States'] = df['States'].str.title()

        # get the totals
        total_revenue = df['Revenue'].sum()
        total_tax = df['Tax'].sum()

        # export the df (for testing purposes)
        # pivot_totals.to_csv(
        #     r"C:\Users\taygu\Downloads\tax templates - sql evan report\monthly.csv".replace("\\", "/")
        #     , index=False)

    except Exception as e:
        logging.error(f"There was an unexpected error with wrangling your data.\n{e}")
        exit()

    # _________________________________________________________________________________________________________________
    # CREATE EXCEL SHEET WITH PROPER FORMATTING FOR ACCOUNTING OFFICE
    # _________________________________________________________________________________________________________________

    try:
        wb = xl.Workbook()
        ws = wb.active

        # report headers
        ws['A1'] = 'Account'
        ws['A2'] = 'Report'
        ws['A3'] = 'Month'
        ws['A4'] = ''
        ws['B1'] = account_name
        ws['B2'] = 'Revenue/Tax breakdown by state'
        ws['B3'] = year_month_properly_formatted
        [ws.merge_cells(cells) for cells in ['B1:C1', 'B2:C2', 'B3:C3']]  # merge top cells

        # add column headers of the df into ws (states, revenue and tax)
        for index, row in enumerate(df.columns, 1):
            ws.cell(row=5, column=index, value=row)

        # add values from df into ws
        for index, row in enumerate(df.values, 6):
            for col_index, value in enumerate(row, 1):
                ws.cell(row=index, column=col_index, value=value)

        # add grand totals row (create it first by adding +1 last row, else it will overwrite the last row/state)
        ws[f"A{ws.max_row +1}"] = 'Total'  # only +1 to first col in new row, else it will create new row for each col
        ws[f"B{ws.max_row}"] = total_revenue
        ws[f"C{ws.max_row}"] = total_tax

        # formatting as currency
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=3):
            for cell in row:
                ws[cell.coordinate].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'

        # center align the rows (create separate loop bc the currency format one starts at row 6 but we need entire col)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = xl.styles.Alignment(horizontal='center', vertical='center')

        # space out the columns
        for col in range(1, 4):
            col_letter = xl.utils.get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20

        # make bold necessary cells
        cells_to_bold = ['A1', 'A2', 'A3', 'A5', 'B5', 'C5', f'A{ws.max_row}', f'B{ws.max_row}', f'C{ws.max_row}']
        for cell in cells_to_bold:
            ws[cell].font = xl.styles.Font(bold=True)

        # highlight necessary cells
        cells_to_highlight = ['A5', 'B5', 'C5', f'A{ws.max_row}', f'B{ws.max_row}', f'C{ws.max_row}']
        for cell in cells_to_highlight:
            ws[cell].fill = xl.styles.PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        # add filter for headers
        ws.auto_filter.ref = f'A5:C{ws.max_row}'

        # save to .xlsx workbook
        wb.save(f"{tax_files_dir}/{account_name} - Revenue Tax breakdown - {year_month_properly_formatted}.xlsx")

    except Exception as e:
        logging.error(f"There was an error with formatting your workbook with openpyxl.\n{e}")
        exit()

    print(f"Successfully generated report for account '{account_name}' for {year_month_properly_formatted}.")


if __name__ == '__main__':
    main()
