import io
import logging
from calendar import month_name, monthrange
from datetime import datetime
from typing import Tuple

import openpyxl as xl
import pandas as pd

from Utilities.config import full_account_names
from Utilities.report_tools import ReportDownloadOrchestrator
from Utilities.utils import BlobHandler, DateRanges, DateRangeOption, Style, states


class TaxRevenueReportGenerator:
    """
    Generates a tax/revenue report by-state out of Amazon API orders data and stores it as .xlsx in your blob container
    
    Parameters:
        -`account_name` (str): The initials of the account - must match the key-vault & env-var names, (e.g. 'po')
        -`start_date` (str): The starting date of the report (format='%m-%d-%Y')
        -`end_date` (str): The end date of the report (format='%m-%d-%Y')
            
    Considerations:
        -Pass None for `start_date`, `end_date` dates to return a report for the entire previous month  
        
        -Eager-loads the `ReportDownloadOrchestrator` and `GenerateFBAReport` classes - will raise error if your 
         Azure account, key vault, storage account, and environment variables are missing or not configured
    """
    def __init__(self, account_name: str, start_date: str, end_date: str):
        # validate date query params
        self.date_ranges = DateRanges()
        self.start_date = start_date
        self.end_date = end_date
        self.start_date, self.end_date = self._validate_date_inputs(self.start_date, self.end_date)

        # split date range passed into 31 day intervals 
        self.report_ranges = self.date_ranges.date_range_splitter(self.start_date, self.end_date)
        
        # name attributes
        self.account_name = account_name
        self.account_name_full = self._set_account_name_full()  # relies on config.full_account_names dict
        
        # class utils
        self.default_dates_used = False
        self.date_name = None
        self.df = None
        self.report_name = None
        
        # validate credentials, env-vars, get keys, access token 
        self.orchestrator = ReportDownloadOrchestrator(self.account_name)
        
        # storage account client (not eager-loading - too heavy, saving for method where used)
        self.blobclient = None

    def _reset_attributes(self) -> None:
        """Resets instance attributes - clears state without having to manually reassign attributes"""
        self.default_dates_used = False
        self.start_date = None
        self.end_date = None
        self.date_name = None
        self.df = None
        self.report_name = None    

    def _validate_date_inputs(self, start_date: str, end_date: str) -> Tuple[str, str]:        
        """Validates `start_date` and `end_date` passed to class instance, formats to '%m-%d-%Y')"""  
          
        # default to previous months range if start or end dates are omitted
        if start_date is None or end_date is None:
            logging.warning(
                "Detected a missing date query parameter, defaulting to report for the previous month. "
                "If this was done in error, specify BOTH 'start_date' and 'end_date' URI parameters and try again"
                )
            start_date, end_date = self.date_ranges.set_default_date_range(default_to=DateRangeOption.PREVIOUS_MONTH)
            self.default_dates_used = True
                    
        # basic validation of date params        
        self.date_ranges.validate_date_logic(start_date=start_date, end_date=end_date)

        # clean inputs in case they are deformed or passed improperly 
        start_date, end_date = self.date_ranges.clean_date_input(start_date=start_date, end_date=end_date)

        logging.info(f"Date inputs validated - proceeding with report for date range: {start_date} - {end_date}")                
        return start_date, end_date 
    
    @staticmethod
    def _no_sales_pivot_table_creator() -> pd.DataFrame:
        """Returns a df with columns=['State', 'Revenue', 'Tax'], with each state containing $0 revenue and tax"""
        state_names = list(states.values())  # see Utils.states dict for reference 
        no_sales_df = pd.DataFrame(state_names + ['Total'], columns=['States'])
        no_sales_df = no_sales_df.assign(Revenue=0, Tax=0)
        return no_sales_df
    
    def _set_account_name_full(self) -> str:
        """Populates `account_name_full` attribute by pulling the full name from config.full_account_names dict"""
        try: 
            return full_account_names[self.account_name.lower()]
        
        except Exception as _:
            logging.warning(
                f"No full name for '{self.account_name}' found in config.full_account_names, proceeding as-is"
                )
            return self.account_name.upper()
 
    def _set_date_name(self) -> str:
        """
        Creates a formal date name for report title and updates the `date_name` attribute with the result
        
        Examples:
            -If the start/end dates constitute a full month (e.g. '11-01-2024 - 11-30-2024), returns 'November 2024' 
                        
            -If 'uneven' dates are provided (e.g. start_date='9/15/2024' & end_date='10/07/2024'), the 
             date_name attribute returned will be populated simply as '09-15-2024 - 10-07-2024'
        """
        # if user left start/end dates blank, default to entire previous month as range
        if self.default_dates_used is True:
            previous_month = self.date_ranges.last_months_date_range[0].split("-")[0]
            previous_year = self.date_ranges.last_months_date_range[0].split("-")[-1]
            self.date_name = f"{month_name[int(previous_month)]} {previous_year}"
            
        # if user provided date parameters...
        else:
            # if start_date and end_date equate to an entire month, return monthname_year
            start_date_input = datetime.strptime(self.start_date, '%m-%d-%Y')
            end_date_input = datetime.strptime(self.end_date, '%m-%d-%Y')
            first_date_of_that_month = start_date_input.replace(day=1)
            last_day_of_that_month = monthrange(start_date_input.year, start_date_input.month)[1]
            last_date_of_that_month = start_date_input.replace(day=last_day_of_that_month) 
            
            if all([
                start_date_input == first_date_of_that_month,
                end_date_input == last_date_of_that_month,
                start_date_input.month == end_date_input.month,
                start_date_input.year == end_date_input.year
            ]):
                self.date_name = f"{month_name[start_date_input.month]} {start_date_input.year}"

            # if its not an entire month, or in any other case, just return the range as a str            
            else:
                self.date_name = f"{self.start_date} - {self.end_date}"                    
        
        return self.date_name
    
    def get_orders_data(self, start_date: str, end_date: str, max_retries: int = 10) -> str:
        """
        Requests/downloads FBA orders data from the 'Reports' SP-API for the date range specified 

        Parameters:
            -`start_date` (str): The start date of the range you wish to generate orders for (format="%m-%d-%Y")
            -`end_date` (str): The end date of the range you wish to generate orders for (format="%m-%d-%Y")
            -`max_retries` (int): n times to check status with exponential backup before passing a failure (default=10)

        Returns:
            -(str): Tabular orders data, saved as json str (also updates `df` instance attribute)\n       
            -Updates the `start_date`, `end_date`, `default_dates_used` and `df` instance attributes

        Considerations:
            -Default action: pass None to `start_date`/`end_date` to return report for the entire previous month\n
            
            -The API can only generate up to 31 days per request. Thus, for longer ranges, it will take some time
             to complete             
        """                        
        try:
            order_data_json = self.orchestrator.get_report(
                report_type='GET_FLAT_FILE_ALL_ORDERS_DATA_BY_ORDER_DATE_GENERAL',
                start_date=start_date,
                end_date=end_date,
                max_retries=max_retries
                )
            
        # adding generic except-block just for the log - underlying 'get_report' method handles all errors/retries  
        except Exception as e :
            logging.error(f"Could not generate order data for dates {start_date} - {end_date}: {str(e)}")
            raise
        
        # 'get_report' method returns tabular data in json, must convert back into Pandas df
        try:
            self.df = pd.read_json(io.StringIO(order_data_json))
            return self.df
        
        except Exception as e:
            logging.info("Failure in 'get_orders_data' method - could not parse json data into pd.df")
            raise RuntimeError(f"Could not parse orders data json to Pandas DataFrame: {str(e)}")
        
    def tax_report_compiler(self, df: pd.DataFrame) -> pd.DataFrame:
        """Generates a pivot-table pd.df out of the `GET_FLAT_FILE_ALL_ORDERS_DATA_BY_ORDER_DATE_GENERAL` report
        
        Parameters:
            -`df` (pd.DataFrame): the raw `GET_FLAT_FILE_ALL_ORDERS_DATA_BY_ORDER_DATE_GENERAL` report
        
        Returns:
            -(pd.DataFrame): pivot-table with total revenue & tax aggregated by state. Returns all states, 
            even if it had 0 sales, columns=['State', 'Revenue', 'Tax']
        
        Considerations;
            -This report handles all US states and District of Columbia (51 in total, +1 'Totals' row)
            
            -Overseas US bases are excluded, but can be added by appending the `Utilities.utils.states` dict
            with an entry that contains "OTHER" as the value (key can be anything)
        """        
        if not isinstance(df, pd.DataFrame):
            raise TypeError("Please provide a Pandas DataFrame for the 'df' parameter input")
        
        #  must faux-populate df attribute if passing a df from local source without first running 'get_orders_data'
        if self.df is None:
            logging.warning(
                "No 'df' attribute located. It is likely you are attempting to compile a tax report before running "
                "'get_orders_data' first - if you are intentionally passing a df from a different source, ignore this "
                "warning (Note: any date query parameters you passed to the URI will be ignored - start/end dates "
                "will be inferred from the orders data)"
                )
            self.df = df

        # if empty (no sales), return a pivot table anyway, with $0 revenue/tax for each state
        if df.empty:
            logging.warning("Warning: no sales for this date range, generating blank report")            
            return self._no_sales_pivot_table_creator()
        
        # basic df validation
        required_columns = [
            'ship-country', 
            'item-status', 
            'product-name',
            'purchase-date',
            'item-price',
            'item-tax',
            'ship-state'
        ]
        for column in required_columns:
            if column not in df.columns:
                raise KeyError(f"Could not locate required column '{column}' in your DataFrame")
                
        # format iso column to date object
        df['purchase-date-formatted'] = pd.to_datetime(df['purchase-date'].str.split("T").str[0]).dt.date        

        # set new start/end dates (must set start/end dates to None to infer dates from the orders data) 
        if self.start_date is None and self.end_date is None:
            self.start_date = df['purchase-date-formatted'].min().strftime("%m-%d-%Y")
            self.end_date = df['purchase-date-formatted'].max().strftime("%m-%d-%Y")
            self.default_dates_used = False
        
        # convert the date attributes back to date obj, in order to be able to compare with df dates 
        start_date_input = datetime.strptime(self.start_date, "%m-%d-%Y").date()
        end_date_input = datetime.strptime(self.end_date, "%m-%d-%Y").date()
        
        # apply filters (drop duplicates, US purchases only, no cancelled orders, no removal orders, date filter)
        df = df.drop_duplicates() \
            .loc[
                lambda x:
                    (x['ship-country'] == 'US') &
                    (x['item-status'] != 'Cancelled') &
                    (x['product-name'] != '-') & 
                    (x['purchase-date-formatted'] >= start_date_input) & 
                    (x['purchase-date-formatted'] <= end_date_input)
                ]  # a few out-of-range dates always make it through in these reports, need to filter them out

        # if df is not empty yet the data doesn't contain anything meaningful, return the no sales pivot table
        # note: this line must go after the date checks above, else it will terminate before setting the new 
        # start/end date attributes, and the date portion of the report name wont populate correctly
        if df['item-price'].sum() == 0 and df['item-tax'].sum() == 0:
            return self._no_sales_pivot_table_creator()

        # clean states data and normalize the names using Utilities.utils.states
        df = df.assign(ship_state_FIXED = df['ship-state'].str.upper().str.replace(".", "").replace(states))

        # map any non-state US locations (overseas territories, army bases) into an "OTHER" bin
        df['ship_state_FIXED'] = df['ship_state_FIXED'].map(lambda x: 'OTHER' if x not in states.values() else x)

        # pivot by revenue/tax
        df_pivot = df.groupby('ship_state_FIXED').agg({'item-price': 'sum', 'item-tax': 'sum'})

        # join back to the states dictionary (need full list of all states, even if they have 0 sales)
        states_df = pd.DataFrame(states.values()).rename(columns={0:'ship_state_FIXED'})
        final_df = pd.merge(states_df, df_pivot, on='ship_state_FIXED', how='left').fillna(0)

        # final clean
        final_df.rename(
            columns={'ship_state_FIXED': 'States', 'item-price': 'Revenue', 'item-tax': 'Tax'}, 
            inplace=True
        )
        final_df['States'] = final_df['States'].str.title()

        # create a 'Total' row, append to the bottom of main df
        totals_row = pd.DataFrame({
            'States': 'Total', 
            'Revenue': final_df['Revenue'].sum(), 
            'Tax': final_df['Tax'].sum()
            },
            index=[0]
        )
        final_df = pd.concat([final_df, totals_row], ignore_index=True)
        
        return final_df

    def tax_report_formatter(self, df: pd.DataFrame) -> io.BytesIO:
        """
        Creates a visually formatted .xlsx workbook out of the pivot-table DataFrame 
        
        Parameters:
            -`df` (pd.DataFrame): a pivoted DataFrame with 'States', 'Revenue' and 'Tax' columns
        
        Returns:
            (io.BytesIO): in-memory object containing the finished report        
        """
        # basic validation         
        if not isinstance(df, pd.DataFrame):
            raise TypeError("Please provide a Pandas DataFrame for the 'df' parameter input")

        required_columns = ['States', 'Revenue', 'Tax']
        for column in required_columns:
            if column not in df.columns:
                raise KeyError(
                    f"Could not locate required column '{column}' in your DataFrame. "
                    f"Make sure you ran the 'tax_report_compiler' first"
                    )
                
        # load wb, styler pen
        wb = xl.Workbook()
        ws = wb.active
        styler = Style(ws)

        # this is required to format cell 'B3', but may be populated already if ran 'set_tax_report_name' already    
        if self.date_name is None:
            self._set_date_name()
        
        # build report headers
        ws['A1'] = 'Account'
        ws['A2'] = 'Report'
        ws['A3'] = 'Month'
        ws['A4'] = ''
        ws['B1'] = self.account_name_full
        ws['B2'] = 'Revenue/Tax breakdown by state'
        ws['B3'] = self.date_name
        [ws.merge_cells(cells) for cells in ['B1:C1', 'B2:C2', 'B3:C3']]  # merge top cells

        # add df headers into ws (e.g. States, Revenue and Tax)
        for index, row in enumerate(df.columns, 1):
            ws.cell(row=5, column=index, value=row)

        # add df contents into ws
        for index, row in enumerate(df.values, 6):
            for col_index, value in enumerate(row, 1):
                ws.cell(row=index, column=col_index, value=value)

        # formatting Revenue/Tax cols as currency
        [styler.currency_formatter(columns=col, min_row=6, currency=True) for col in ['B', 'C']]

        # center/align everything
        styler.align_and_center(padding=4)

        # make necessary cells bold
        cells_to_bold = ['A1', 'A2', 'A3', 'A5', 'B5', 'C5', f'A{ws.max_row}', f'B{ws.max_row}', f'C{ws.max_row}']
        [styler.apply_styles_to_cell(cell, bold=True, highlighter=False) for cell in cells_to_bold]

        # highlight necessary cells
        cells_to_color = ['A5', 'B5', 'C5', f'A{ws.max_row}', f'B{ws.max_row}', f'C{ws.max_row}']
        [styler.apply_styles_to_cell(cell, bold=False, highlighter=True, color='DDEBF7') for cell in cells_to_color]

        # add filter for headers
        ws.auto_filter.ref = f'A5:C{ws.max_row}'

        # save to buffer
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer

    def set_tax_report_name(self) -> str:
        """
        Generates a full report name for saving/uploading the final report
        
        Returns: 
            -(str) in the format `<full account name> - Revenue Tax Breakdown - <formal date name>` 

        Considerations:
            -.xlsx extension is not included in the output 
        """
        # basic validation
        if self.df is None:
            raise ValueError("Attribute 'df' is unpopulated - please provide a tax report before naming it")

        # this could be populated already, if you ran the formatter method first
        if self.date_name is None:
            self._set_date_name()

        self.report_name = f"{self.account_name_full} - Revenue Tax Breakdown - {self.date_name}"
        return self.report_name
    
    def generate_tax_report(self, storage_account: str, container_name: str, max_retries: int = 10) -> None:
        """
        Orchestrates the entire tax report automation workflow:\n
            Requests FBA orders data from SP-API for the date range passed to the class instance\n
            Pivots the orders data by state\n
            Visually formats into finished report\n
            Names the report\n
            Uploads as .xlsx to blob container
        
        Parameters:
            -`storage_account` (str): The name of your Azure storage account
            -`container_name` (str): The name of the blob container in your storage account
            -`max_retries` (int): n times to check status with exp. backup before passing a failure (Default=10)
        
        Considerations:
            -Orders data is requested in 31 day chunks. For longer date-ranges, job will take a while to complete
        """
        # instantiate a blob client
        if self.blobclient is None:
            try:
                self.blobclient = BlobHandler(storage_account=storage_account, container_name=container_name)
            except Exception as e:
                logging.error(f"Could not start blob client: {str(e)}")
                raise
        
        # iterate through date ranges list and generate orders data in 31D blocks, then concat to one final df 
        df = []
        min_max_dates = []  # get the first and last dates in here to be able to name the report properly
        n_ranges = len(self.report_ranges) 
        for index, date_range in enumerate(self.report_ranges):
            try:
                start_date, end_date = date_range
                if index == 0:
                    min_max_dates.append(start_date)
                if index == n_ranges-1:
                    min_max_dates.append(end_date)

                _df = self.get_orders_data(start_date=start_date, end_date=end_date, max_retries=max_retries)
                df.append(_df)                
            
                logging.info(f"Retrieved orders data for '{self.account_name}' for {start_date} - {end_date}")
            
            except Exception as e:
                logging.error(f"Failed getting orders data from SP-API for {start_date} - {end_date}: {str(e)}")
                raise         
        
        # convert list of dfs into one df
        try:                  
            df = pd.concat(df, ignore_index=True).drop_duplicates()  # dates may overlap, drop any doubles
        except Exception as e:
            logging.error(f"Could not concatenate dfs: {str(e)}")
            
        # pivot the data    
        try:
            pivot_df = self.tax_report_compiler(df=df)
            logging.info(f"Successfully assembled pivot table")
        except Exception as e:
            logging.error(f"Could not compile orders data: {str(e)}")
            raise

        # format the data
        try:
            xlsx_file = self.tax_report_formatter(df=pivot_df)
            logging.info("Successfully formatted pivot table and saved to buffer as .xlsx")
        except Exception as e:
            logging.error(f"Could not format the pivot table data: {str(e)}")
            raise
        
        # generate name
        try:
            name = self.set_tax_report_name()
            logging.info("Successfully set report name")
        except Exception as e:
            logging.error(f"Could not set the report name: {str(e)}")
            raise
        
        # upload to azure
        try:
            self.blobclient.save_to_blob(xlsx_file, f"{name}.xlsx")  # this generates an info log on its own
        except Exception as e:
            logging.error(f"Could not upload to blob client {str(e)}")
            raise
    
    def generate_tax_report_from_local(self, df: pd.DataFrame) -> Tuple[str, io.BytesIO]:
        """
        Generates a tax report without calling the API or using `get_orders_data`:\n
            Pivots the orders data by state\n
            Visually formats into finished report\n
            Names the report\n
            Returns a finished .xlsx file as a memory object 
        
        Parameters:
            -`df` (pd.DataFrame): The orders data `GET_FLAT_FILE_ALL_ORDERS_DATA_BY_ORDER_DATE_GENERAL`
            -`max_retries` (int): n times to check status with exponential backup before raising error (Default=10)
        
        Returns:
            - (Tuple[str, io.BytesIO]) the name of the report in position 0, and the report itself in position 1
        
        Considerations:
            -This method infers the start/end dates through the orders data, and disregards/overwrites the existing
            `start_date` and `end_date` attributes 
            
            -Unlike the `generate_tax_report` method, the output of this method is not saved to a blob container. It
            returns a name and buffer object - you can unpack the Tuple and do with it whatever you may need
            
            -Since you aren't calling to SP-API or Azure, the `account_name` parameters only real use in this method 
            is to title the report
        
        Example:
            >>report_generator = TaxRevenueReportGenerator(account_name='test', start_date=None, end_date=None)\n
            >>local_df = pd.read_csv("file.csv")\n
            >>name, output = report_generator.generate_tax_report_from_local(df=local_df)\n            
            >>wb = openpyxl.load_workbook(output)\n
            >>wb.save(f"{name}.xlsx")\n              
        """        
        # note: no validation needed here, everything is handled in 'tax_report_compiler'

        # overwrite the start/end attributes (else the report will use the date query params)
        self._reset_attributes()
        
        # from here, its the same as with the API method...
        
        # compile report
        try:
            pivot_df = self.tax_report_compiler(df=df)
        except Exception as e:
            logging.error(f"Could not compile orders data: {str(e)}")
            raise
        
        # format report
        try:
            output = self.tax_report_formatter(df=pivot_df)
        except Exception as e:
            logging.error(f"Could not format the pivot table data: {str(e)}")
            raise
        
        # generate name
        try:
            name = self.set_tax_report_name()
        except Exception as e:
            logging.error(f"Could not set the report name: {str(e)}")
            raise
        
        # return output as tuple
        logging.info(f"Successfully generated tax report from local df: '{name}'")
        return name, output  # leave it to the user to decide how/where to save the report
